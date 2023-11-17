import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager


def headless_chrome():
    from selenium.webdriver.chrome.service import Service
    ops = webdriver.ChromeOptions()
    ops.add_argument('--headless')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=ops)
    return driver


driver = headless_chrome()
#driver.get("https://www.tridenthotels.com/site-map")
print(driver.title)
# wb = load_workbook("D:\Automation\PyAutomation\search.xlsx")
# ws = wb.active
# rows=ws.max_row
# cols=ws.max_column
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(ws.cell(r,c).value)

def getrowCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return (ws.max_row)



def getColumnCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return (ws.max_column)

def readData(file,sheetName,rowno,colno):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return ws.cell(rowno,colno).value

def writeData(file,sheetName,rowno,colno,data):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    ws.cell(rowno, colno).value=data
    wb.save(file)

def fillRedColor(file,sheetName,rowno,colno):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    greenfill=PatternFill(start_color='68b212',end_color='68b212',fill_type='Solid')
    ws.cell(rowno,colno).fill=greenfill
    wb.save(file)

def fillGreenColor(file,sheetName,rowno,colno):
    wb = openpyxl.load_workbook(file)
    ws=wb[sheetName]
    redfill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='Solid')
    ws.cell(rowno,colno).fill=redfill
    wb.save(file)